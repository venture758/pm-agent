from __future__ import annotations

import io
import json
import tempfile
import unittest
from pathlib import Path

from mysql_test_utils import ensure_mysql_schema, get_test_database_url, reset_mysql_test_data
from openpyxl import Workbook

from pm_agent.api import create_api_app
from pm_agent.api_service import WorkspaceService
from pm_agent.models import AssignmentRecommendation, MemberProfile, ModuleKnowledgeEntry, RequirementItem
from pm_agent.story_excel_import import CANONICAL_STORY_HEADERS


class FakeKnowledgeLlm:
    def __init__(self, *, parsed: dict | None = None, chat_error: Exception | None = None, parse_error: Exception | None = None) -> None:
        self.parsed = parsed or {}
        self.chat_error = chat_error
        self.parse_error = parse_error

    def chat_completion(self, messages=None, **_kwargs) -> str:
        if self.chat_error is not None:
            raise self.chat_error
        return json.dumps(self.parsed, ensure_ascii=False)

    def parse_json_response(self, text: str) -> dict:
        if self.parse_error is not None:
            raise self.parse_error
        return json.loads(text)


@unittest.skipUnless(get_test_database_url(), "PM_AGENT_TEST_DATABASE_URL 未配置")
class WebWorkbenchTest(unittest.TestCase):
    def setUp(self) -> None:
        self.database_url = get_test_database_url()
        assert self.database_url is not None
        self.store = ensure_mysql_schema(self.database_url)
        reset_mysql_test_data(self.store)

    def _create_service(self, _root: str | Path) -> WorkspaceService:
        return WorkspaceService(database_url=self.database_url)

    def _create_app(self, _root: str | Path):
        return create_api_app(database_url=self.database_url)

    def _prepare_confirmable_workspace(self, service: WorkspaceService) -> None:
        service.create_managed_member(
            "default",
            {
                "name": "李祥",
                "role": "developer",
                "skills": "发票",
                "experience": "高",
                "workload": 0.2,
                "capacity": 1.0,
            },
        )
        service.create_module_entry(
            "default",
            {
                "big_module": "税务",
                "function_module": "发票接口",
                "primary_owner": "李祥",
                "familiar_members": ["李祥"],
            },
        )
        service.update_draft(
            "default",
            {"message_text": "1. 发票查验接口替换 https://example.com/1 优先级高"},
        )
        service.generate_recommendations("default")

    def _create_module_workbook(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "2026泾渭云后端技术角度的业务模块"
        sheet.append(["主要负责人", "B角", "大模块", "功能模块", "李祥", "王海林"])
        sheet.append(["", "", "", "", "李祥", "王海林"])
        sheet.append(["李祥", "王海林", "税务", "发票接口", "熟悉", "了解"])
        workbook.save(path)

    def _create_story_workbook(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "列表数据"
        sheet.append(["序号", "用户故事名称", "状态", "用户故事编码", "计划提测完成时间", "负责人", "测试人员名称", "优先级", "关联缺陷", "详细说明（URL）", "开发人员名称"])
        sheet.append([1, "发票需求", "开发中", "PRJ-1", "2026-04-01", "李祥", "余萍", "高", 1, "https://example.com/story", "李祥"])
        sheet.append(["合计"])
        workbook.save(path)

    def _create_story_full_workbook(
        self,
        path: Path,
        rows: list[dict[str, object]] | None = None,
        *,
        headers: list[str] | None = None,
    ) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "列表数据"
        effective_headers = headers or CANONICAL_STORY_HEADERS
        sheet.append(effective_headers)
        rows = rows or []
        for row_data in rows:
            sheet.append([row_data.get(header) for header in effective_headers])
        workbook.save(path)

    def _create_task_workbook(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "列表数据"
        sheet.append(["序号", "任务编号", "关联用户故事", "任务名称", "任务类型", "负责人", "状态", "预计开始时间", "预计结束时间", "计划人天", "实际人天", "缺陷总数"])
        sheet.append([1, "TK-1", "发票需求", "设计与编码【发票需求】", "设计与编码", "李祥", "进行中", "2026-04-01", "2026-04-02", 1.0, 0.0, 0])
        workbook.save(path)

    def _request(self, app, method: str, path: str, payload: dict | None = None):
        body = json.dumps(payload or {}).encode("utf-8") if payload is not None else b""
        result: dict[str, object] = {}
        path_info, _, query_string = path.partition("?")

        def start_response(status, headers):
            result["status"] = status
            result["headers"] = headers

        response = b"".join(
            app(
                {
                    "REQUEST_METHOD": method,
                    "PATH_INFO": path_info,
                    "QUERY_STRING": query_string,
                    "CONTENT_LENGTH": str(len(body)),
                    "CONTENT_TYPE": "application/json",
                    "wsgi.input": io.BytesIO(body),
                },
                start_response,
            )
        )
        return result["status"], json.loads(response.decode("utf-8"))

    def _multipart_request(self, app, method: str, path: str, files: dict[str, tuple[str, bytes]]):
        boundary = "----pm-agent-boundary"
        body = bytearray()
        for field_name, (filename, content) in files.items():
            body.extend(f"--{boundary}\r\n".encode("utf-8"))
            body.extend(
                (
                    f'Content-Disposition: form-data; name="{field_name}"; '
                    f'filename="{filename}"\r\n'
                ).encode("utf-8")
            )
            body.extend(b"Content-Type: application/octet-stream\r\n\r\n")
            body.extend(content)
            body.extend(b"\r\n")
        body.extend(f"--{boundary}--\r\n".encode("utf-8"))

        result: dict[str, object] = {}

        def start_response(status, headers):
            result["status"] = status
            result["headers"] = headers

        path_info, _, query_string = path.partition("?")
        response = b"".join(
            app(
                {
                    "REQUEST_METHOD": method,
                    "PATH_INFO": path_info,
                    "QUERY_STRING": query_string,
                    "CONTENT_LENGTH": str(len(body)),
                    "CONTENT_TYPE": f"multipart/form-data; boundary={boundary}",
                    "wsgi.input": io.BytesIO(bytes(body)),
                },
                start_response,
            )
        )
        return result["status"], json.loads(response.decode("utf-8"))

    def test_database_configuration_is_required(self) -> None:
        with self.assertRaisesRegex(ValueError, "PM_AGENT_DATABASE_URL"):
            WorkspaceService()
        with self.assertRaisesRegex(ValueError, "PM_AGENT_DATABASE_URL"):
            create_api_app()

    def test_service_flow(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            service = self._create_service(root)
            service.create_managed_member(
                "default",
                {
                    "name": "李祥",
                    "role": "developer",
                    "skills": "发票,接口",
                    "experience": "高",
                    "workload": 0.2,
                    "capacity": 1.0,
                },
            )
            service.create_managed_member(
                "default",
                {
                    "name": "余萍",
                    "role": "tester",
                    "skills": "测试",
                    "experience": "中",
                    "workload": 0.1,
                    "capacity": 1.0,
                },
            )
            service.create_managed_member(
                "default",
                {
                    "name": "王海林",
                    "role": "developer",
                    "skills": "税务",
                    "experience": "中",
                    "workload": 0.1,
                    "capacity": 1.0,
                },
            )
            service.create_module_entry(
                "default",
                {
                    "big_module": "税务",
                    "function_module": "发票接口",
                    "primary_owner": "李祥",
                    "backup_owners": ["余萍"],
                    "familiar_members": ["李祥"],
                    "aware_members": ["余萍"],
                },
            )
            service.update_draft(
                "default",
                {
                    "message_text": "1. 发票查验接口替换 https://example.com/1 优先级高",
                },
            )
            workspace_after_draft = service.workspaces.load_workspace("default")
            self.assertEqual(["1"], workspace_after_draft.current_session_requirement_ids)
            module_page = service.list_module_entries("default")
            self.assertEqual(50, module_page["module_page"]["page_size"])
            self.assertEqual(1, module_page["module_page"]["total"])
            payload = service.generate_recommendations("default")
            self.assertEqual(1, len(payload["recommendations"]))
            with self.store.connection() as connection:
                cursor = connection.cursor()
                cursor.execute(
                    "SELECT COUNT(1) FROM workspace_recommendations WHERE workspace_id = %s",
                    ("default",),
                )
                recommendation_count = cursor.fetchone()[0]
            self.assertEqual(1, recommendation_count)

            workspace_before_confirm = service.workspaces.load_workspace("default")
            session_id_before_confirm = workspace_before_confirm.current_session_id
            confirmed = service.confirm_assignments("default", {"1": {"action": "accept"}})
            self.assertEqual(1, len(confirmed["confirmed_assignments"]))
            self.assertEqual([], confirmed["handoff"]["stories"])
            self.assertEqual([], confirmed["handoff"]["tasks"])
            self.assertIn("已记录确认结果（未进入平台同步）", confirmed["messages"][0])
            self.assertEqual(0, len(confirmed["recommendations"]))
            with self.store.connection() as connection:
                cursor = connection.cursor()
                cursor.execute(
                    (
                        "SELECT session_id, confirmed_count, payload_json "
                        "FROM workspace_confirmation_records WHERE workspace_id = %s "
                        "ORDER BY id DESC LIMIT 1"
                    ),
                    ("default",),
                )
                record_row = cursor.fetchone()
            self.assertIsNotNone(record_row)
            self.assertEqual(session_id_before_confirm, record_row[0])
            self.assertEqual(1, record_row[1])
            payload_json = json.loads(record_row[2])
            self.assertEqual("default", payload_json["workspace_id"])
            self.assertEqual(session_id_before_confirm, payload_json["session_id"])
            self.assertEqual(1, payload_json["confirmed_count"])
            self.assertEqual(1, len(payload_json["confirmed_assignments"]))
            self.assertEqual("1", payload_json["confirmed_assignments"][0]["requirement_id"])
            workspace_after_confirm = service.workspaces.load_workspace("default")
            self.assertEqual([], workspace_after_confirm.current_session_requirement_ids)

            module_path = root / "module.xlsx"
            story_path = root / "story.xlsx"
            task_path = root / "task.xlsx"
            self._create_module_workbook(module_path)
            self._create_story_workbook(story_path)
            self._create_task_workbook(task_path)

            module_payload = service.upload_module_knowledge("default", "module.xlsx", module_path.read_bytes())
            self.assertGreater(module_payload["knowledge_base_summary"]["entry_count"], 0)
            self.assertEqual("database", module_payload["uploads"][-1]["storage_backend"])
            self.assertEqual(len(module_path.read_bytes()), module_payload["uploads"][-1]["file_size"])
            self.assertFalse((root / "workspaces" / "default" / "uploads").exists())

            sync_payload = service.sync_platform_files(
                "default",
                ("story.xlsx", story_path.read_bytes()),
                ("task.xlsx", task_path.read_bytes()),
            )
            self.assertIsNotNone(sync_payload["latest_sync_batch"])
            self.assertFalse((root / "workspaces" / "default" / "uploads").exists())

            alerts = service.get_monitoring("default")
            self.assertTrue(alerts["alerts"])

            insights = service.get_insights("default")
            self.assertIn("heatmap", insights["insights"])

    def test_api_routes(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            app = self._create_app(tmpdir)
            status, payload = self._request(
                app,
                "POST",
                "/api/workspaces/default/members",
                {
                    "name": "李祥",
                    "role": "developer",
                    "skills": "发票,接口",
                    "experience": "高",
                    "workload": 0.2,
                    "capacity": 1.0,
                },
            )
            self.assertTrue(status.startswith("200"))
            self.assertEqual(1, len(payload["managed_members"]))

            status, payload = self._request(
                app,
                "POST",
                "/api/workspaces/default/modules",
                {
                    "big_module": "税务",
                    "function_module": "发票接口",
                    "primary_owner": "李祥",
                    "familiar_members": ["李祥"],
                },
            )
            self.assertTrue(status.startswith("200"))
            self.assertEqual(1, len(payload["module_entries"]))

            status, payload = self._request(
                app,
                "POST",
                "/api/workspaces/default/modules",
                {
                    "big_module": "支付",
                    "function_module": "对账",
                    "primary_owner": "李祥",
                    "familiar_members": ["李祥"],
                },
            )
            self.assertTrue(status.startswith("200"))
            self.assertEqual(2, len(payload["module_entries"]))

            status, payload = self._request(
                app,
                "GET",
                "/api/workspaces/default/modules?page=1&page_size=1&big_module=税&function_module=发票&primary_owner=李",
            )
            self.assertTrue(status.startswith("200"))
            self.assertEqual(1, len(payload["module_entries"]))
            self.assertEqual("税务", payload["module_entries"][0]["big_module"])
            self.assertEqual(1, payload["module_page"]["page"])
            self.assertEqual(1, payload["module_page"]["page_size"])
            self.assertEqual(1, payload["module_page"]["total"])
            self.assertEqual("税", payload["module_page"]["filters"]["big_module"])

            status, payload = self._request(
                app,
                "PUT",
                "/api/workspaces/default/draft",
                {
                    "message_text": "1. 发票查验接口替换 https://example.com/1 优先级高",
                },
            )
            self.assertTrue(status.startswith("200"))
            self.assertEqual("default", payload["workspace_id"])

            status, payload = self._request(app, "POST", "/api/workspaces/default/recommendations", {})
            self.assertTrue(status.startswith("200"))
            self.assertEqual(1, len(payload["recommendations"]))

            status, payload = self._request(app, "DELETE", "/api/workspaces/default/recommendations/1")
            self.assertTrue(status.startswith("200"))
            self.assertEqual(0, len(payload["recommendations"]))
            status, payload = self._request(app, "DELETE", "/api/workspaces/default/recommendations/1")
            self.assertTrue(status.startswith("200"))
            self.assertEqual(0, len(payload["recommendations"]))

            status, payload = self._request(app, "POST", "/api/workspaces/default/recommendations", {})
            self.assertTrue(status.startswith("200"))
            self.assertEqual(1, len(payload["recommendations"]))

            status, payload = self._request(
                app,
                "POST",
                "/api/workspaces/default/recommendations/batch-delete",
                {"requirement_ids": ["1"]},
            )
            self.assertTrue(status.startswith("200"))
            self.assertEqual(0, len(payload["recommendations"]))

            status, payload = self._request(app, "POST", "/api/workspaces/default/recommendations", {})
            self.assertTrue(status.startswith("200"))
            self.assertEqual(1, len(payload["recommendations"]))

            status, payload = self._request(
                app,
                "POST",
                "/api/workspaces/default/confirmations",
                {"actions": {"1": {"action": "accept"}}},
            )
            self.assertTrue(status.startswith("200"))
            self.assertEqual(1, len(payload["confirmed_assignments"]))
            self.assertEqual(0, len(payload["recommendations"]))
            self.assertEqual([], payload["handoff"]["stories"])
            self.assertEqual([], payload["handoff"]["tasks"])
            self.assertIn("已记录确认结果（未进入平台同步）", payload["messages"][0])
            session_id = payload["latest_knowledge_update"]["session_id"]
            status, detail_payload = self._request(
                app,
                "GET",
                f"/api/workspaces/default/confirmations/{session_id}/requirements/1/knowledge-update-modules",
            )
            self.assertTrue(status.startswith("200"))
            self.assertEqual(1, detail_payload["total"])
            self.assertEqual("税务::发票接口", detail_payload["items"][0]["module_key"])

    def test_story_only_upload_service_supports_partial_fail_and_upsert(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            service = self._create_service(root)

            def story_row(code: str | None, name: str, seq: int) -> dict[str, object]:
                row = {header: None for header in CANONICAL_STORY_HEADERS}
                row["序号"] = seq
                row["用户故事编码"] = code
                row["用户故事名称"] = name
                row["状态"] = "计划"
                row["负责人"] = "李祥"
                row["测试人员名称"] = "余萍"
                row["优先级"] = "高"
                row["计划人天"] = 2
                row["计划开发人天"] = 1.5
                row["关联缺陷"] = 0
                row["详细说明（URL）"] = "https://example.com/story"
                row["开发人员名称"] = "李祥;王海林"
                return row

            story_path = root / "story-full.xlsx"
            self._create_story_full_workbook(
                story_path,
                rows=[
                    story_row("PRJ-00731813", "发票查验接口替换", 1),
                    story_row(None, "缺少编码的故事", 2),
                ],
            )

            payload = service.upload_story_file("default", ("story-full.xlsx", story_path.read_bytes()))
            summary = payload["latest_story_import"]
            self.assertEqual(2, summary["total"])
            self.assertEqual(1, summary["created"])
            self.assertEqual(0, summary["updated"])
            self.assertEqual(1, summary["failed"])
            self.assertEqual(1, len(summary["errors"]))
            self.assertEqual(1, len(payload["handoff"]["stories"]))
            self.assertEqual("PRJ-00731813", payload["handoff"]["stories"][0]["user_story_code"])

            self._create_story_full_workbook(
                story_path,
                rows=[story_row("PRJ-00731813", "发票查验接口替换-更新", 1)],
            )
            payload2 = service.upload_story_file("default", ("story-full.xlsx", story_path.read_bytes()))
            summary2 = payload2["latest_story_import"]
            self.assertEqual(1, summary2["total"])
            self.assertEqual(0, summary2["created"])
            self.assertEqual(1, summary2["updated"])
            self.assertEqual(0, summary2["failed"])
            self.assertEqual("发票查验接口替换-更新", payload2["handoff"]["stories"][0]["name"])

    def test_story_only_upload_api_success_and_full_fail(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            app = self._create_app(tmpdir)
            root = Path(tmpdir)

            valid_path = root / "story-valid.xlsx"
            valid_row = {header: None for header in CANONICAL_STORY_HEADERS}
            valid_row["序号"] = 1
            valid_row["用户故事编码"] = "PRJ-00730001"
            valid_row["用户故事名称"] = "品牌官网优化"
            valid_row["状态"] = "计划"
            valid_row["优先级"] = "中"
            self._create_story_full_workbook(valid_path, rows=[valid_row])
            ok_status, ok_payload = self._multipart_request(
                app,
                "POST",
                "/api/workspaces/default/uploads/story-only",
                {"story_file": ("story-valid.xlsx", valid_path.read_bytes())},
            )
            self.assertTrue(ok_status.startswith("200"))
            self.assertEqual(0, ok_payload["latest_story_import"]["failed"])
            self.assertEqual(1, len(ok_payload["handoff"]["stories"]))

            invalid_path = root / "story-invalid.xlsx"
            invalid_headers = [header for header in CANONICAL_STORY_HEADERS if header != "用户故事编码"]
            invalid_row = {header: None for header in invalid_headers}
            invalid_row["用户故事名称"] = "缺少编码字段"
            self._create_story_full_workbook(
                invalid_path,
                rows=[invalid_row],
                headers=invalid_headers,
            )
            bad_status, bad_payload = self._multipart_request(
                app,
                "POST",
                "/api/workspaces/default/uploads/story-only",
                {"story_file": ("story-invalid.xlsx", invalid_path.read_bytes())},
            )
            self.assertTrue(bad_status.startswith("400"))
            self.assertIn("缺失列", bad_payload["error"])

    def test_delete_recommendation_normalizes_requirement_id(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            service = self._create_service(tmpdir)
            workspace = service.workspaces.load_workspace("default")
            workspace.recommendations = [
                AssignmentRecommendation(
                    requirement_id="2.0",
                    title="示例需求",
                    development_owner="张三",
                )
            ]
            service.workspaces.save_workspace(workspace)

            payload = service.delete_recommendation("default", "2")
            self.assertEqual(0, len(payload["recommendations"]))

    def test_confirm_assignments_persists_internal_record_only(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            service = self._create_service(root)
            service.create_managed_member(
                "default",
                {
                    "name": "李祥",
                    "role": "developer",
                    "skills": "发票",
                    "experience": "高",
                    "workload": 0.2,
                    "capacity": 1.0,
                },
            )
            service.create_module_entry(
                "default",
                {
                    "big_module": "税务",
                    "function_module": "发票接口",
                    "primary_owner": "李祥",
                    "familiar_members": ["李祥"],
                },
            )
            service.update_draft(
                "default",
                {"message_text": "1. 发票查验接口替换 https://example.com/1 优先级高"},
            )
            service.generate_recommendations("default")
            session_id = service.workspaces.load_workspace("default").current_session_id

            payload = service.confirm_assignments("default", {"1": {"action": "accept"}})
            self.assertEqual([], payload["handoff"]["stories"])
            self.assertEqual([], payload["handoff"]["tasks"])
            self.assertIn("已记录确认结果（未进入平台同步）", payload["messages"][0])

            with self.store.connection() as connection:
                cursor = connection.cursor()
                cursor.execute(
                    "SELECT workspace_id, session_id, confirmed_count, payload_json "
                    "FROM workspace_confirmation_records WHERE workspace_id = %s",
                    ("default",),
                )
                rows = cursor.fetchall()

            self.assertEqual(1, len(rows))
            workspace_id, row_session_id, confirmed_count, payload_json = rows[0]
            self.assertEqual("default", workspace_id)
            self.assertEqual(session_id, row_session_id)
            self.assertEqual(1, confirmed_count)
            record = json.loads(payload_json)
            self.assertEqual("default", record["workspace_id"])
            self.assertEqual(session_id, record["session_id"])
            self.assertEqual(1, record["confirmed_count"])

    def test_confirm_assignments_persists_successful_knowledge_update_record(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            service = self._create_service(root)
            service.agent._llm = FakeKnowledgeLlm(
                parsed={
                    "reply": "建议培养发票接口 B 角",
                    "knowledge_updates": {
                        "suggested_familiarity": [{"member": "余萍", "to": "了解"}],
                        "suggested_modules": [],
                    },
                    "optimization_suggestions": [{"type": "single_point", "suggestion": "培养 B 角"}],
                }
            )
            self._prepare_confirmable_workspace(service)
            session_id = service.workspaces.load_workspace("default").current_session_id

            payload = service.confirm_assignments("default", {"1": {"action": "accept"}})

            self.assertEqual("success", payload["latest_knowledge_update"]["status"])
            self.assertEqual(session_id, payload["latest_knowledge_update"]["session_id"])
            self.assertEqual("建议培养发票接口 B 角", payload["latest_knowledge_update"]["reply"])
            self.assertTrue(payload["latest_knowledge_update"]["has_module_diff_records"])
            self.assertEqual(1, payload["latest_knowledge_update"]["module_change_count"])
            self.assertEqual(1, payload["latest_knowledge_update"]["requirement_change_count"])
            self.assertEqual(1, len(payload["latest_knowledge_update"]["module_diff_records"]))

            with self.store.connection() as connection:
                cursor = connection.cursor()
                cursor.execute(
                    "SELECT session_id, status, payload_json FROM workspace_knowledge_update_records WHERE workspace_id = %s",
                    ("default",),
                )
                rows = cursor.fetchall()
                cursor.execute(
                    "SELECT requirement_id, module_key, changed, payload_json "
                    "FROM workspace_knowledge_update_module_diff_records WHERE workspace_id = %s",
                    ("default",),
                )
                module_rows = cursor.fetchall()

            self.assertEqual(1, len(rows))
            row_session_id, status, payload_json = rows[0]
            self.assertEqual(session_id, row_session_id)
            self.assertEqual("success", status)
            record = json.loads(payload_json)
            self.assertEqual("success", record["status"])
            self.assertEqual("建议培养发票接口 B 角", record["reply"])
            self.assertEqual(1, len(module_rows))
            requirement_id, module_key, changed, module_payload_json = module_rows[0]
            self.assertEqual("1", requirement_id)
            self.assertEqual("税务::发票接口", module_key)
            self.assertEqual(1, changed)
            module_record = json.loads(module_payload_json)
            self.assertEqual("1", module_record["requirement_id"])
            self.assertIn("before_snapshot", module_record)
            self.assertIn("after_snapshot", module_record)
            self.assertGreaterEqual(module_record["diff_summary"]["changed_field_count"], 1)

            history = service.list_confirmation_records("default", page=1, page_size=20)
            self.assertEqual("success", history["items"][0]["knowledge_update"]["status"])
            detail = service.get_knowledge_update_module_diff_records("default", session_id, "1")
            self.assertEqual(1, detail["total"])
            self.assertEqual("税务::发票接口", detail["items"][0]["module_key"])

    def test_confirm_assignments_skips_knowledge_update_when_llm_unavailable(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            service = self._create_service(tmpdir)
            self._prepare_confirmable_workspace(service)

            payload = service.confirm_assignments("default", {"1": {"action": "accept"}})

            self.assertEqual(1, len(payload["confirmed_assignments"]))
            self.assertEqual("skipped", payload["latest_knowledge_update"]["status"])
            self.assertEqual("LLM 未配置", payload["latest_knowledge_update"]["reply"])
            self.assertTrue(payload["latest_knowledge_update"]["has_module_diff_records"])
            history = service.list_confirmation_records("default", page=1, page_size=20)
            self.assertEqual("skipped", history["items"][0]["knowledge_update"]["status"])

    def test_confirm_assignments_marks_failed_knowledge_update_without_rollback(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            service = self._create_service(tmpdir)
            service.agent._llm = FakeKnowledgeLlm(parse_error=ValueError("无法解析 JSON"))
            self._prepare_confirmable_workspace(service)

            payload = service.confirm_assignments("default", {"1": {"action": "accept"}})

            self.assertEqual(1, len(payload["confirmed_assignments"]))
            self.assertEqual("failed", payload["latest_knowledge_update"]["status"])
            self.assertIn("无法解析 JSON", payload["latest_knowledge_update"]["error_message"])
            workspace = service.workspaces.load_workspace("default")
            self.assertEqual(1, len(workspace.confirmed_assignments))
            self.assertEqual("failed", workspace.latest_knowledge_update.status)
            self.assertTrue(workspace.latest_knowledge_update.has_module_diff_records)

    def test_module_diff_records_can_mark_unchanged_snapshot(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            service = self._create_service(tmpdir)
            entry = ModuleKnowledgeEntry(
                key="税务::发票接口",
                big_module="税务",
                function_module="发票接口",
                primary_owner="李祥",
                familiar_members=["李祥"],
            )
            before_snapshot = service._module_entry_snapshot(entry)
            changed, diff_summary = service._build_module_snapshot_diff(before_snapshot, before_snapshot)
            self.assertFalse(changed)
            self.assertEqual(0, diff_summary["changed_field_count"])
            self.assertEqual([], diff_summary["modified_fields"])

    def test_chat_recommendation_flow(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            service = self._create_service(tmpdir)
            service.create_managed_member(
                "default",
                {
                    "name": "李祥",
                    "role": "developer",
                    "skills": "发票,接口",
                    "experience": "高",
                    "workload": 0.2,
                    "capacity": 1.0,
                },
            )
            service.create_module_entry(
                "default",
                {
                    "big_module": "税务",
                    "function_module": "发票接口",
                    "primary_owner": "李祥",
                    "familiar_members": ["李祥"],
                },
            )

            def fake_parse(_message: str, _members, task_history=None, task_records=None):
                requirements = [
                    RequirementItem(
                        requirement_id="CHAT-1",
                        title="发票查验接口替换",
                        priority="高",
                        risk="中",
                        complexity="中",
                        skills=["发票", "接口"],
                    )
                ]
                return requirements, "已解析 1 条需求"

            service.agent.parse_requirements_with_llm = fake_parse
            chat_payload = service.send_chat_message(
                "default",
                {"message": "1. 发票查验接口替换 https://example.com/1 优先级高"},
            )
            self.assertEqual(2, len(chat_payload["draft"]["chat_messages"]))
            self.assertEqual(1, len(chat_payload["requirements"]))

            recommendation_payload = service.generate_recommendations("default")
            self.assertEqual(1, len(recommendation_payload["recommendations"]))

    def test_generate_recommendations_only_current_session(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            service = self._create_service(tmpdir)
            service.create_managed_member(
                "default",
                {
                    "name": "李祥",
                    "role": "developer",
                    "skills": "发票,接口",
                    "experience": "高",
                    "workload": 0.2,
                    "capacity": 1.0,
                },
            )
            service.create_module_entry(
                "default",
                {
                    "big_module": "税务",
                    "function_module": "发票接口",
                    "primary_owner": "李祥",
                    "familiar_members": ["李祥"],
                },
            )

            service.update_draft(
                "default",
                {"message_text": "1. 历史需求 https://example.com/old 优先级高"},
            )
            first_payload = service.generate_recommendations("default")
            self.assertEqual(["1"], [item["requirement_id"] for item in first_payload["recommendations"]])

            service.update_draft(
                "default",
                {"message_text": "2. 当前需求 https://example.com/new 优先级高"},
            )
            second_payload = service.generate_recommendations("default")
            self.assertEqual(["2"], [item["requirement_id"] for item in second_payload["recommendations"]])

            workspace = service.workspaces.load_workspace("default")
            self.assertEqual(["2"], workspace.current_session_requirement_ids)
            self.assertEqual(
                {"1", "2"},
                {item.requirement_id for item in workspace.normalized_requirements},
            )

    def test_api_recommendation_requires_current_session(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            app = self._create_app(tmpdir)
            status, payload = self._request(
                app,
                "POST",
                "/api/workspaces/default/members",
                {
                    "name": "李祥",
                    "role": "developer",
                    "skills": "发票",
                    "experience": "高",
                    "workload": 0.2,
                    "capacity": 1.0,
                },
            )
            self.assertTrue(status.startswith("200"))

            status, payload = self._request(
                app,
                "POST",
                "/api/workspaces/default/modules",
                {
                    "big_module": "税务",
                    "function_module": "发票接口",
                    "primary_owner": "李祥",
                    "familiar_members": ["李祥"],
                },
            )
            self.assertTrue(status.startswith("200"))

            status, payload = self._request(
                app,
                "POST",
                "/api/workspaces/default/recommendations",
                {},
            )
            self.assertTrue(status.startswith("400"))
            self.assertIn("当前会话没有可推荐需求", payload["error"])

    def test_maintenance_validation(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            service = self._create_service(tmpdir)
            with self.assertRaisesRegex(ValueError, "请先在人员管理页面维护成员"):
                service.create_module_entry(
                    "default",
                    {
                        "big_module": "税务",
                        "function_module": "发票接口",
                        "primary_owner": "李祥",
                    },
                )
            service.create_managed_member(
                "default",
                {
                    "name": "李祥",
                    "role": "developer",
                    "skills": "发票",
                    "experience": "高",
                    "workload": 0.2,
                    "capacity": 1.0,
                },
            )
            service.create_module_entry(
                "default",
                {
                    "big_module": "税务",
                    "function_module": "发票接口",
                    "primary_owner": "李祥",
                    "familiar_members": ["李祥"],
                },
            )

            with self.assertRaisesRegex(ValueError, "模块键已存在"):
                service.create_module_entry(
                    "default",
                    {
                        "big_module": "税务",
                        "function_module": "发票接口",
                        "primary_owner": "李祥",
                        "familiar_members": ["李祥"],
                    },
                )

            with self.assertRaisesRegex(ValueError, "B角未在人员管理中维护"):
                service.create_module_entry(
                    "default",
                    {
                        "big_module": "税务",
                        "function_module": "发票对账",
                        "primary_owner": "李祥",
                        "backup_owners": ["不存在成员"],
                    },
                )

            with self.assertRaisesRegex(ValueError, "成员容量必须大于 0"):
                service.create_managed_member(
                    "default",
                    {
                        "name": "王海林",
                        "role": "developer",
                        "skills": "后端",
                        "experience": "中",
                        "workload": 0.1,
                        "capacity": 0,
                    },
                )

    def test_update_module_entry_supports_stale_path_key(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            service = self._create_service(tmpdir)
            service.create_managed_member(
                "default",
                {
                    "name": "李祥",
                    "role": "developer",
                    "skills": "发票",
                    "experience": "高",
                    "workload": 0.2,
                    "capacity": 1.0,
                },
            )
            service.create_module_entry(
                "default",
                {
                    "big_module": "税务",
                    "function_module": "发票接口",
                    "primary_owner": "李祥",
                    "familiar_members": ["李祥"],
                },
            )

            service.update_module_entry(
                "default",
                "税务::发票接口",
                {
                    "big_module": "税务-新",
                    "function_module": "发票接口",
                    "primary_owner": "李祥",
                    "familiar_members": ["李祥"],
                    "original_key": "税务::发票接口",
                },
            )
            payload = service.update_module_entry(
                "default",
                "税务::发票接口",
                {
                    "big_module": "税务-新",
                    "function_module": "发票接口",
                    "primary_owner": "李祥",
                    "familiar_members": ["李祥"],
                    "original_key": "税务::发票接口",
                },
            )
            self.assertEqual(1, len(payload["module_entries"]))
            self.assertEqual("税务-新::发票接口", payload["module_entries"][0]["key"])

    def test_local_legacy_files_are_ignored(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            legacy_workspace_dir = root / "workspaces" / "default"
            legacy_workspace_dir.mkdir(parents=True, exist_ok=True)

            (root / "state.json").write_text(json.dumps({"messages": ["legacy-agent"]}, ensure_ascii=False), encoding="utf-8")
            (legacy_workspace_dir / "workspace.json").write_text(
                json.dumps({"workspace_id": "default", "title": "默认工作区", "messages": ["legacy"]}, ensure_ascii=False),
                encoding="utf-8",
            )

            service = self._create_service(root)
            payload = service.get_workspace("default")
            self.assertEqual("default", payload["title"])
            self.assertEqual([], payload["messages"])

            with self.store.connection() as connection:
                cursor = connection.cursor()
                cursor.execute("SELECT COUNT(1) FROM workspace_states WHERE workspace_id = %s", ("default",))
                self.assertEqual(0, cursor.fetchone()[0])
                cursor.execute("SELECT COUNT(1) FROM agent_states WHERE state_key = %s", ("global",))
                self.assertEqual(0, cursor.fetchone()[0])

    def test_local_workspace_json_does_not_backfill_member_table(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            legacy_workspace_dir = root / "workspaces" / "default"
            legacy_workspace_dir.mkdir(parents=True, exist_ok=True)

            (legacy_workspace_dir / "workspace.json").write_text(
                json.dumps(
                    {
                        "workspace_id": "default",
                        "title": "默认工作区",
                        "managed_members": [
                            {
                                "name": "李祥",
                                "role": "developer",
                                "skills": ["税务", "接口"],
                                "experience_level": "高",
                                "workload": 0.1,
                                "capacity": 1.0,
                                "constraints": ["仅白天可排期"],
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            service = self._create_service(root)
            payload = service.get_workspace("default")
            self.assertEqual([], payload["managed_members"])

            with self.store.connection() as connection:
                cursor = connection.cursor()
                cursor.execute(
                    "SELECT COUNT(1) FROM workspace_managed_members WHERE workspace_id = %s",
                    ("default",),
                )
                count = cursor.fetchone()[0]
                self.assertEqual(0, count)

    def test_confirmation_history_pagination(self) -> None:
        """GET /api/workspaces/:id/confirmations should return paginated records."""
        with tempfile.TemporaryDirectory() as tmpdir:
            app = self._create_app(tmpdir)

            # Confirm a record first
            status, _ = self._request(
                app, "POST", "/api/workspaces/default/members",
                {"name": "张三", "role": "developer", "skills": "Java", "experience": "高", "workload": 0.0, "capacity": 1.0},
            )
            self.assertTrue(status.startswith("200"))

            # Create a recommendation and confirm
            status, payload = self._request(app, "POST", "/api/workspaces/default/recommendations", {})
            # Need to set up session first; use a simpler approach via API
            # Just verify the endpoint returns empty correctly
            status, payload = self._request(app, "GET", "/api/workspaces/default/confirmations")
            self.assertTrue(status.startswith("200"))
            self.assertEqual(0, payload["total"])
            self.assertEqual([], payload["items"])
            self.assertEqual(1, payload["page"])
            self.assertEqual(20, payload["page_size"])

    def test_v2_prefix_confirmation_history_pagination(self) -> None:
        """GET /api/v2/workspaces/:id/confirmations should be backward compatible."""
        with tempfile.TemporaryDirectory() as tmpdir:
            app = self._create_app(tmpdir)
            status, payload = self._request(app, "GET", "/api/v2/workspaces/default/confirmations")
            self.assertTrue(status.startswith("200"))
            self.assertEqual(0, payload["total"])
            self.assertEqual([], payload["items"])
            self.assertEqual(1, payload["page"])
            self.assertEqual(20, payload["page_size"])

    def test_v2_prefix_health(self) -> None:
        """GET /api/v2/health should map to health endpoint."""
        with tempfile.TemporaryDirectory() as tmpdir:
            app = self._create_app(tmpdir)
            status, payload = self._request(app, "GET", "/api/v2/health")
            self.assertTrue(status.startswith("200"))
            self.assertEqual("ok", payload["status"])

    def test_confirmation_history_returns_records(self) -> None:
        """After confirming assignments, GET confirmations should return the record."""
        with tempfile.TemporaryDirectory() as tmpdir:
            service = self._create_service(tmpdir)

            # Set up members and modules
            service.agent.state.module_entries = {
                "税务::发票接口": ModuleKnowledgeEntry(
                    key="税务::发票接口", big_module="税务", function_module="发票接口",
                    primary_owner="李祥", backup_owners=["王海林"],
                ),
            }
            service.agent.save()

            # Create workspace with a recommendation
            workspace = service.workspaces.load_workspace("default")
            workspace.member_profiles = [
                MemberProfile(name="李祥", workload=0.2, capacity=1.0),
            ]
            workspace.managed_members = [
                MemberProfile(name="李祥", workload=0.2, capacity=1.0),
            ]
            workspace.module_entries = list(service.agent.state.module_entries.values())
            workspace.current_session_id = "session-test"
            workspace.current_session_requirement_ids = ["1"]
            workspace.normalized_requirements = [
                RequirementItem(requirement_id="1", title="测试需求", matched_module_keys=["税务::发票接口"]),
            ]
            service.workspaces.save_workspace(workspace)

            # Generate recommendations
            rec_payload = service.generate_recommendations("default")
            self.assertGreater(len(rec_payload["recommendations"]), 0)

            # Confirm
            conf_payload = service.confirm_assignments("default", {})
            self.assertGreater(len(conf_payload["confirmed_assignments"]), 0)

            # Query confirmation history
            history = service.list_confirmation_records("default", page=1, page_size=20)
            self.assertEqual(1, history["total"])
            self.assertEqual(1, len(history["items"]))
            self.assertEqual(1, history["items"][0]["confirmed_count"])

    def test_confirmation_history_out_of_range_page(self) -> None:
        """Requesting a page beyond total should return empty items with correct total."""
        with tempfile.TemporaryDirectory() as tmpdir:
            service = self._create_service(tmpdir)
            items, total = service.workspaces.load_confirmation_records("default", page=5, page_size=10)
            self.assertEqual(0, total)
            self.assertEqual([], items)

    def test_story_pagination_default(self) -> None:
        """GET /api/workspaces/:id/stories returns paginated response with default page=1, pageSize=20."""
        with tempfile.TemporaryDirectory() as tmpdir:
            app = self._create_app(tmpdir)
            # Insert story records directly via database
            store = app.service.workspaces
            with store.database.connection() as conn:
                cursor = conn.cursor()
                for i in range(25):
                    cursor.execute(
                        "INSERT INTO workspace_story_records ("
                        "workspace_id, user_story_code, user_story_name, modified_time, imported_at, updated_at"
                        ") VALUES (%s, %s, %s, %s, %s, %s)",
                        ("ws1", f"US-{i:03d}", f"Story {i}", f"2026-01-{(i % 28) + 1:02d}", "2026-01-01", "2026-01-01"),
                    )
            status, payload = self._request(app, "GET", "/api/workspaces/ws1/stories")
            self.assertTrue(status.startswith("200"))
            self.assertEqual(25, payload["total"])
            self.assertEqual(1, payload["page"])
            self.assertEqual(20, payload["page_size"])
            self.assertEqual(2, payload["total_pages"])
            self.assertEqual(20, len(payload["items"]))

    def test_story_pagination_keyword_filter(self) -> None:
        """Keyword filter on stories endpoint returns matching records."""
        with tempfile.TemporaryDirectory() as tmpdir:
            app = self._create_app(tmpdir)
            store = app.service.workspaces
            with store.database.connection() as conn:
                cursor = conn.cursor()
                for i in range(10):
                    cursor.execute(
                        "INSERT INTO workspace_story_records ("
                        "workspace_id, user_story_code, user_story_name, modified_time, imported_at, updated_at"
                        ") VALUES (%s, %s, %s, %s, %s, %s)",
                        ("ws1", f"US-{i:03d}", f"Story {i}", "2026-01-01", "2026-01-01", "2026-01-01"),
                    )
            status, payload = self._request(app, "GET", "/api/workspaces/ws1/stories?keyword=US-005")
            self.assertTrue(status.startswith("200"))
            self.assertEqual(1, payload["total"])
            self.assertEqual(1, len(payload["items"]))

    def test_story_pagination_empty_workspace(self) -> None:
        """Empty workspace returns zero total."""
        with tempfile.TemporaryDirectory() as tmpdir:
            app = self._create_app(tmpdir)
            status, payload = self._request(app, "GET", "/api/workspaces/empty-ws/stories")
            self.assertTrue(status.startswith("200"))
            self.assertEqual(0, payload["total"])
            self.assertEqual(1, payload["total_pages"])
            self.assertEqual([], payload["items"])

    def test_story_pagination_out_of_range(self) -> None:
        """Out-of-range page returns empty items."""
        with tempfile.TemporaryDirectory() as tmpdir:
            app = self._create_app(tmpdir)
            store = app.service.workspaces
            with store.database.connection() as conn:
                cursor = conn.cursor()
                for i in range(5):
                    cursor.execute(
                        "INSERT INTO workspace_story_records ("
                        "workspace_id, user_story_code, user_story_name, modified_time, imported_at, updated_at"
                        ") VALUES (%s, %s, %s, %s, %s, %s)",
                        ("ws1", f"US-{i:03d}", f"Story {i}", "2026-01-01", "2026-01-01", "2026-01-01"),
                    )
            status, payload = self._request(app, "GET", "/api/workspaces/ws1/stories?page=999&pageSize=20")
            self.assertTrue(status.startswith("200"))
            self.assertEqual(5, payload["total"])
            self.assertEqual([], payload["items"])

    def test_task_pagination_returns_paginated_response(self) -> None:
        """GET /api/workspaces/:id/tasks with page/pageSize returns paginated response."""
        with tempfile.TemporaryDirectory() as tmpdir:
            app = self._create_app(tmpdir)
            # Insert task records directly via database
            store = app.service.workspaces
            with store.database.connection() as conn:
                cursor = conn.cursor()
                for i in range(25):
                    owner = "李祥" if i < 10 else "王海林"
                    status = "进行中" if i < 15 else "已完成"
                    cursor.execute(
                        "INSERT INTO workspace_task_records ("
                        "workspace_id, task_code, name, owner, status, project_name, "
                        "modified_time, imported_at, updated_at"
                        ") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
                        ("ws1", f"TK-{i:03d}", f"Task {i}", owner, status, f"项目{i % 3}",
                         "2026-01-01", "2026-01-01", "2026-01-01"),
                    )

            status, payload = self._request(app, "GET", "/api/workspaces/ws1/tasks?page=1&pageSize=10")
            self.assertTrue(status.startswith("200"))
            self.assertIn("items", payload)
            self.assertEqual(25, payload["total"])
            self.assertEqual(1, payload["page"])
            self.assertEqual(10, payload["page_size"])
            self.assertEqual(3, payload["total_pages"])
            self.assertEqual(10, len(payload["items"]))

    def test_task_pagination_with_filters(self) -> None:
        """Task pagination combined with owner/status filters."""
        with tempfile.TemporaryDirectory() as tmpdir:
            app = self._create_app(tmpdir)
            store = app.service.workspaces
            with store.database.connection() as conn:
                cursor = conn.cursor()
                for i in range(15):
                    owner = "李祥" if i < 5 else ("王海林" if i < 10 else "余萍")
                    status = "进行中" if i < 8 else "已完成"
                    cursor.execute(
                        "INSERT INTO workspace_task_records ("
                        "workspace_id, task_code, name, owner, status, project_name, "
                        "modified_time, imported_at, updated_at"
                        ") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
                        ("ws1", f"TK-{i:03d}", f"Task {i}", owner, status, "项目A",
                         "2026-01-01", "2026-01-01", "2026-01-01"),
                    )

            status, payload = self._request(
                app, "GET",
                "/api/workspaces/ws1/tasks?page=1&pageSize=10&owner=李祥&status=进行中",
            )
            self.assertTrue(status.startswith("200"))
            self.assertEqual(5, payload["total"])
            self.assertEqual(5, len(payload["items"]))


if __name__ == "__main__":
    unittest.main()
