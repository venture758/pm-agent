from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import re
from typing import Any, Mapping

from openpyxl import load_workbook

from .excel_io import ExcelInput, workbook_stream
from .utils import normalize_text, split_names

CANONICAL_TASK_HEADERS: list[str] = [
    "序号",
    "任务编号",
    "关联用户故事",
    "任务名称",
    "任务类型",
    "负责人",
    "状态",
    "预计开始时间",
    "预计结束时间",
    "备注",
    "完成时间",
    "计划人天",
    "实际人天",
    "产品",
    "模块路径",
    "所属项目",
    "版本",
    "迭代阶段",
    "项目组",
    "参与人",
    "客户名称",
    "缺陷总数",
    "关联代码",
    "创建人",
    "创建时间",
    "修改人",
    "修改时间",
]

# Canonical header -> DB column name
TASK_HEADER_COLUMN_MAP: dict[str, str] = {
    "序号": "sequence_no",
    "任务编号": "task_code",
    "关联用户故事": "related_story",
    "任务名称": "name",
    "任务类型": "task_type",
    "负责人": "owner",
    "状态": "status",
    "预计开始时间": "estimated_start",
    "预计结束时间": "estimated_end",
    "备注": "remark",
    "完成时间": "completed_time",
    "计划人天": "planned_person_days",
    "实际人天": "actual_person_days",
    "产品": "product",
    "模块路径": "module_path",
    "所属项目": "project_name",
    "版本": "version",
    "迭代阶段": "iteration_phase",
    "项目组": "project_group",
    "参与人": "participants",
    "客户名称": "customer_name",
    "缺陷总数": "defect_count",
    "关联代码": "related_code",
    "创建人": "created_by",
    "创建时间": "created_time",
    "修改人": "modified_by",
    "修改时间": "modified_time",
}

TASK_DB_COLUMNS: list[str] = [TASK_HEADER_COLUMN_MAP[header] for header in CANONICAL_TASK_HEADERS]

HEADER_ALIASES: dict[str, str] = {
    "关联用户故事": "关联用户故事",
    "预计开始": "预计开始时间",
    "预计结束": "预计结束时间",
}

NUMERIC_HEADERS = {
    "序号",
    "计划人天",
    "实际人天",
    "缺陷总数",
}

DATETIME_HEADERS = {
    "预计开始时间",
    "预计结束时间",
    "完成时间",
    "创建时间",
    "修改时间",
}


@dataclass
class TaskImportError:
    row: int
    reason: str


@dataclass
class TaskExcelImportResult:
    total_rows: int
    valid_rows: list[dict[str, Any]]
    errors: list[TaskImportError]
    missing_headers: list[str]
    unknown_headers: list[str]


class TaskExcelHeaderError(ValueError):
    def __init__(self, missing_headers: list[str], unknown_headers: list[str]) -> None:
        missing_text = "、".join(missing_headers) if missing_headers else "无"
        unknown_text = "、".join(unknown_headers) if unknown_headers else "无"
        super().__init__(f"任务 Excel 模板不匹配，缺失列: {missing_text}；未知列: {unknown_text}")
        self.missing_headers = missing_headers
        self.unknown_headers = unknown_headers


def normalize_task_header(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = text.replace("（", "(").replace("）", ")").replace("：", ":")
    text = re.sub(r"\s+", "", text)
    return text.lower()


def _resolve_canonical_header(value: Any) -> str:
    normalized = normalize_task_header(value)
    if not normalized:
        return ""
    # Check aliases first
    for alias, canonical in HEADER_ALIASES.items():
        if normalize_task_header(alias) == normalized:
            return canonical
    # Match canonical headers
    for header in CANONICAL_TASK_HEADERS:
        if normalize_task_header(header) == normalized:
            return header
    return ""


def _parse_numeric(value: Any) -> float | None:
    text = normalize_text(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_datetime(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    text = normalize_text(value)
    if not text:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt in {"%Y-%m-%d", "%Y/%m/%d"}:
                return parsed.date().isoformat()
            return parsed.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return text


def _parse_task_cell(header: str, value: Any) -> Any:
    if header in NUMERIC_HEADERS:
        return _parse_numeric(value)
    if header in DATETIME_HEADERS:
        return _parse_datetime(value)
    text = normalize_text(value)
    return text or None


def import_task_excel(source: ExcelInput) -> TaskExcelImportResult:
    workbook = load_workbook(workbook_stream(source), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    unknown_headers: list[str] = []
    header_indexes: dict[str, int] = {}

    for idx, raw_header in enumerate(header_row):
        canonical = _resolve_canonical_header(raw_header)
        if not canonical:
            text = normalize_text(raw_header)
            if text:
                unknown_headers.append(text)
            continue
        if canonical not in header_indexes:
            header_indexes[canonical] = idx

    missing_headers = [header for header in CANONICAL_TASK_HEADERS if header not in header_indexes]
    if missing_headers:
        raise TaskExcelHeaderError(missing_headers=missing_headers, unknown_headers=unknown_headers)

    total_rows = 0
    valid_rows: list[dict[str, Any]] = []
    errors: list[TaskImportError] = []

    for row_idx, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        first_value = normalize_text(values[0]) if values else ""
        if first_value == "合计":
            break
        if not any(value not in (None, "") for value in values):
            continue

        total_rows += 1
        parsed_row: dict[str, Any] = {}
        for header in CANONICAL_TASK_HEADERS:
            value = values[header_indexes[header]] if header_indexes[header] < len(values) else None
            parsed_row[TASK_HEADER_COLUMN_MAP[header]] = _parse_task_cell(header, value)

        task_code = normalize_text(parsed_row.get("task_code"))
        if not task_code:
            errors.append(TaskImportError(row=row_idx, reason="缺少任务编号"))
            continue

        parsed_row["task_code"] = task_code
        parsed_row["participants"] = split_names(parsed_row.get("participants"))
        valid_rows.append(parsed_row)

    return TaskExcelImportResult(
        total_rows=total_rows,
        valid_rows=valid_rows,
        errors=errors,
        missing_headers=missing_headers,
        unknown_headers=unknown_headers,
    )


def row_to_task_record(row: Mapping[str, Any]) -> dict[str, Any]:
    """将数据库行转换为 TaskRecord 兼容的字典结构。"""
    participants = row.get("participants")
    if isinstance(participants, str):
        participants = split_names(participants)
    elif participants is None:
        participants = []

    return {
        "task_code": normalize_text(row.get("task_code")),
        "related_story": normalize_text(row.get("related_story")),
        "name": normalize_text(row.get("name")),
        "task_type": normalize_text(row.get("task_type")),
        "owner": normalize_text(row.get("owner")),
        "status": normalize_text(row.get("status")),
        "estimated_start": normalize_text(row.get("estimated_start")),
        "estimated_end": normalize_text(row.get("estimated_end")),
        "remark": normalize_text(row.get("remark")),
        "completed_time": normalize_text(row.get("completed_time")),
        "planned_person_days": float(row.get("planned_person_days") or 0.0),
        "actual_person_days": float(row.get("actual_person_days") or 0.0),
        "product": normalize_text(row.get("product")),
        "module_path": normalize_text(row.get("module_path")),
        "project_name": normalize_text(row.get("project_name")),
        "version": normalize_text(row.get("version")),
        "iteration_phase": normalize_text(row.get("iteration_phase")),
        "project_group": normalize_text(row.get("project_group")),
        "participants": participants,
        "customer_name": normalize_text(row.get("customer_name")),
        "defect_count": int(float(row.get("defect_count") or 0.0)),
        "related_code": normalize_text(row.get("related_code")),
        "created_by": normalize_text(row.get("created_by")),
        "created_time": normalize_text(row.get("created_time")),
        "modified_by": normalize_text(row.get("modified_by")),
        "modified_time": normalize_text(row.get("modified_time")),
        "sequence_no": row.get("sequence_no"),
    }
