from __future__ import annotations

import re
from datetime import datetime
from typing import Iterable, Optional

from openpyxl import load_workbook

from .excel_io import ExcelInput, workbook_stream
from .models import AssignmentHistoryEntry, ConfirmedAssignment, ModuleKnowledgeEntry, RequirementItem
from .utils import familiarity_score, normalize_name, normalize_text, overlap_score, promote_familiarity, split_names, unique_list


def _choose_latest_module_sheet(sheet_names: list[str]) -> str:
    candidates: list[tuple[int, str]] = []
    for sheet_name in sheet_names:
        if "后端技术角度的业务模块" not in sheet_name:
            continue
        match = re.search(r"(20\d{2})", sheet_name)
        year = int(match.group(1)) if match else 0
        candidates.append((year, sheet_name))
    if not candidates:
        raise ValueError("未找到模块划分 sheet")
    return sorted(candidates, reverse=True)[0][1]


def import_module_knowledge_from_excel(source: ExcelInput, imported_at: Optional[str] = None) -> list[ModuleKnowledgeEntry]:
    workbook = load_workbook(workbook_stream(source), read_only=True, data_only=True)
    sheet_name = _choose_latest_module_sheet(workbook.sheetnames)
    sheet = workbook[sheet_name]
    imported_at = imported_at or datetime.utcnow().isoformat()
    member_names = [normalize_text(cell) for cell in next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[4:]]
    current_primary = ""
    current_backup = ""
    current_big_module = ""
    entries: list[ModuleKnowledgeEntry] = []
    year_match = re.search(r"(20\d{2})", sheet_name)
    source_year = int(year_match.group(1)) if year_match else 0
    for row in sheet.iter_rows(min_row=3, values_only=True):
        primary_owner = normalize_text(row[0]) or current_primary
        backup_owner = normalize_text(row[1]) or current_backup
        big_module = normalize_text(row[2]) or current_big_module
        function_module = normalize_text(row[3])
        if not function_module:
            current_primary = primary_owner
            current_backup = backup_owner
            current_big_module = big_module
            continue
        familiarity_by_member: dict[str, str] = {}
        for member, cell in zip(member_names, row[4:4 + len(member_names)]):
            member_name = normalize_text(member)
            familiarity = normalize_text(cell)
            if member_name and familiarity:
                familiarity_by_member[member_name] = familiarity
        key = f"{big_module}::{function_module}"
        entries.append(
            ModuleKnowledgeEntry(
                key=key,
                big_module=big_module,
                function_module=function_module,
                primary_owner=primary_owner,
                backup_owners=split_names(backup_owner),
                familiar_members=[name for name, level in familiarity_by_member.items() if level == "熟悉"],
                aware_members=[
                    name for name, level in familiarity_by_member.items() if level in {"了解", "一般"}
                ],
                unfamiliar_members=[name for name, level in familiarity_by_member.items() if level == "不了解"],
                source_sheet=sheet_name,
                source_year=source_year,
                imported_at=imported_at,
            )
        )
        current_primary = primary_owner
        current_backup = backup_owner
        current_big_module = big_module
    return entries


def merge_module_entries(
    existing: dict[str, ModuleKnowledgeEntry],
    imported_entries: Iterable[ModuleKnowledgeEntry],
) -> dict[str, ModuleKnowledgeEntry]:
    merged = dict(existing)
    for entry in imported_entries:
        old = merged.get(entry.key)
        if old is None:
            merged[entry.key] = entry
            continue
        entry.assignment_history = old.assignment_history
        entry.recent_assignees = old.recent_assignees
        entry.suggested_familiarity = old.suggested_familiarity
        merged[entry.key] = entry
    return merged


def match_requirement_to_modules(
    requirement: RequirementItem,
    module_entries: Iterable[ModuleKnowledgeEntry],
    top_n: int = 3,
) -> list[tuple[ModuleKnowledgeEntry, int]]:
    scored: list[tuple[ModuleKnowledgeEntry, int]] = []
    for entry in module_entries:
        score = overlap_score(entry.function_module, requirement.title) + overlap_score(entry.big_module, requirement.title)
        if score <= 0:
            continue
        if normalize_text(entry.function_module) in normalize_text(requirement.raw_text):
            score += 2
        scored.append((entry, score))
    ranked = sorted(scored, key=lambda item: item[1], reverse=True)
    requirement.matched_module_keys = [entry.key for entry, _ in ranked[:top_n]]
    return ranked[:top_n]


def update_module_knowledge_after_assignment(
    entries: dict[str, ModuleKnowledgeEntry],
    confirmed_assignments: Iterable[ConfirmedAssignment],
) -> dict[str, ModuleKnowledgeEntry]:
    for assignment in confirmed_assignments:
        if not assignment.module_key or assignment.module_key not in entries:
            continue
        entry = entries[assignment.module_key]
        suggested_updates: dict[str, str] = {}
        all_members = [assignment.development_owner, assignment.testing_owner] + assignment.collaborators
        for member in unique_list([member for member in all_members if member]):
            current = entry.familiarity_by_member.get(member) or entry.suggested_familiarity.get(member) or "不了解"
            promoted = promote_familiarity(current, confirmed=True)
            if promoted != current:
                suggested_updates[member] = promoted
                entry.suggested_familiarity[member] = promoted
            if member not in entry.recent_assignees:
                entry.recent_assignees.append(member)
        entry.recent_assignees = entry.recent_assignees[-10:]
        entry.assignment_history.append(
            AssignmentHistoryEntry(
                requirement_id=assignment.requirement_id,
                title=assignment.title,
                development_owner=assignment.development_owner,
                testing_owner=assignment.testing_owner,
                collaborators=assignment.collaborators,
                suggested_familiarity_updates=suggested_updates,
            )
        )
    return entries


def module_entry_coverage(entry: ModuleKnowledgeEntry) -> int:
    return sum(1 for value in entry.familiarity_by_member.values() if familiarity_score(value) >= 2)
