from __future__ import annotations

from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional
from uuid import uuid4

from openpyxl import load_workbook

from .models import AgentState, ImportBatch, StoryRecord, SyncAction, TaskRecord
from .utils import normalize_text, parse_date, safe_float, split_names


def _iter_sheet_rows(path: str | Path) -> tuple[list[str], list[dict[str, object]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [normalize_text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: list[dict[str, object]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = {header: value for header, value in zip(headers, values)}
        first_value = normalize_text(values[0]) if values else ""
        if first_value == "合计":
            break
        if not any(value not in (None, "") for value in values):
            continue
        rows.append(row)
    return headers, rows


def import_story_records(path: str | Path) -> list[StoryRecord]:
    _, rows = _iter_sheet_rows(path)
    stories: list[StoryRecord] = []
    for row in rows:
        code = normalize_text(row.get("用户故事编码"))
        if not code:
            continue
        stories.append(
            StoryRecord(
                user_story_code=code,
                name=normalize_text(row.get("用户故事名称")),
                user_story_name=normalize_text(row.get("用户故事名称")),
                sequence_no=safe_float(row.get("序号"), default=0.0) if normalize_text(row.get("序号")) else None,
                user_story_tag=normalize_text(row.get("用户故事标签")),
                status=normalize_text(row.get("状态")),
                remark=normalize_text(row.get("备注")),
                plan_test_completion_time=normalize_text(row.get("计划提测完成时间")),
                related_task_count=safe_float(row.get("关联任务")),
                owner=normalize_text(row.get("负责人")),
                owner_names=normalize_text(row.get("负责人")),
                tester=normalize_text(row.get("测试人员名称")),
                tester_names=normalize_text(row.get("测试人员名称")),
                requirement_owner_names=normalize_text(row.get("需求人员名称")),
                developers=split_names(row.get("开发人员名称") or row.get("负责人")),
                developer_names=normalize_text(row.get("开发人员名称") or row.get("负责人")),
                priority=normalize_text(row.get("优先级")) or "中",
                absolute_priority=safe_float(row.get("绝对优先级")) if normalize_text(row.get("绝对优先级")) else None,
                plan_test_date=normalize_text(row.get("计划提测完成时间") or row.get("计划主干测试完成时间")),
                planned_person_days=safe_float(row.get("计划开发人天") or row.get("计划人天")),
                planned_dev_person_days=safe_float(row.get("计划开发人天")),
                actual_person_days=safe_float(row.get("实际人天")),
                related_case_count=safe_float(row.get("关联用例")) if normalize_text(row.get("关联用例")) else None,
                plan_trunk_test_completion_time=normalize_text(row.get("计划主干测试完成时间")),
                modified_time=normalize_text(row.get("修改时间")),
                acceptance_criteria=normalize_text(row.get("验收标准")),
                detail_url=normalize_text(row.get("详细说明（URL）")),
                project_status=normalize_text(row.get("立项状态")),
                iteration_phase=normalize_text(row.get("迭代阶段")),
                iteration_goal=normalize_text(row.get("迭代目标")),
                release_plan=normalize_text(row.get("发布计划")),
                release_window=normalize_text(row.get("发布窗口")),
                scrum_team=normalize_text(row.get("Scrum团队")),
                product_group=normalize_text(row.get("产品组")),
                project_name=normalize_text(row.get("所属项目")),
                ksm_or_bug_no=normalize_text(row.get("KSM或BUG编号")),
                story_type=normalize_text(row.get("类型")),
                cloud_name=normalize_text(row.get("所属云")),
                application_name=normalize_text(row.get("所属应用")),
                related_story=normalize_text(row.get("关联故事")),
                related_requirement=normalize_text(row.get("关联需求")),
                completed_time=normalize_text(row.get("完成时间")),
                plan_baseline_test_completion_time=normalize_text(row.get("计划基线测试完成时间")),
                has_upgrade_notice=normalize_text(row.get("是否有升级注意事项")),
                change_type=normalize_text(row.get("变动类型")),
                product=normalize_text(row.get("产品")),
                version=normalize_text(row.get("版本")),
                created_time=normalize_text(row.get("创建时间")),
                created_by=normalize_text(row.get("创建人")),
                defects=int(safe_float(row.get("关联缺陷"))),
                related_defect_count=safe_float(row.get("关联缺陷")) if normalize_text(row.get("关联缺陷")) else None,
                module_path=normalize_text(row.get("所属应用")),
            )
        )
    return stories


def import_task_records(path: str | Path) -> list[TaskRecord]:
    _, rows = _iter_sheet_rows(path)
    tasks: list[TaskRecord] = []
    for row in rows:
        code = normalize_text(row.get("任务编号"))
        if not code:
            continue
        tasks.append(
            TaskRecord(
                task_code=code,
                story_code="",
                story_name=normalize_text(row.get("关联用户故事")),
                name=normalize_text(row.get("任务名称")),
                task_type=normalize_text(row.get("任务类型")),
                owner=normalize_text(row.get("负责人")),
                status=normalize_text(row.get("状态")),
                estimated_start=normalize_text(row.get("预计开始时间")),
                estimated_end=normalize_text(row.get("预计结束时间")),
                planned_person_days=safe_float(row.get("计划人天")),
                actual_person_days=safe_float(row.get("实际人天")),
                participants=split_names(row.get("参与人")),
                product=normalize_text(row.get("产品")),
                version=normalize_text(row.get("版本")),
                module_path=normalize_text(row.get("模块路径")),
                defects=int(safe_float(row.get("缺陷总数"))),
            )
        )
    return tasks


def _changed_fields(current: object, incoming: object) -> list[str]:
    current_dict = asdict(current)
    incoming_dict = asdict(incoming)
    return [key for key, value in incoming_dict.items() if current_dict.get(key) != value]


def upsert_story_records(state: AgentState, stories: Iterable[StoryRecord], batch: ImportBatch) -> None:
    for story in stories:
        action = "create"
        changed_fields = list(asdict(story).keys())
        if story.user_story_code in state.stories:
            changed_fields = _changed_fields(state.stories[story.user_story_code], story)
            action = "update"
        state.stories[story.user_story_code] = story
        batch.actions.append(SyncAction(entity_type="story", record_code=story.user_story_code, action=action, changed_fields=changed_fields))


def upsert_task_records(state: AgentState, tasks: Iterable[TaskRecord], batch: ImportBatch) -> None:
    for task in tasks:
        action = "create"
        changed_fields = list(asdict(task).keys())
        if task.task_code in state.tasks:
            changed_fields = _changed_fields(state.tasks[task.task_code], task)
            action = "update"
        state.tasks[task.task_code] = task
        batch.actions.append(SyncAction(entity_type="task", record_code=task.task_code, action=action, changed_fields=changed_fields))


def sync_platform_exports(
    state: AgentState,
    story_excel_path: str | Path,
    task_excel_path: str | Path,
    imported_at: Optional[str] = None,
) -> ImportBatch:
    imported_at = imported_at or datetime.utcnow().isoformat()
    batch = ImportBatch(
        batch_id=f"batch-{uuid4().hex[:8]}",
        imported_at=imported_at,
        source_files=[str(story_excel_path), str(task_excel_path)],
    )
    upsert_story_records(state, import_story_records(story_excel_path), batch)
    upsert_task_records(state, import_task_records(task_excel_path), batch)
    state.import_batches.append(batch)
    return batch
