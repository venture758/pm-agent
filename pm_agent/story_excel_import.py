from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import re
from typing import Any, Mapping

from openpyxl import load_workbook

from .excel_io import ExcelInput, workbook_stream
from .models import StoryRecord
from .utils import normalize_text, split_names

CANONICAL_STORY_HEADERS: list[str] = [
    "序号",
    "用户故事名称",
    "用户故事标签",
    "状态",
    "用户故事编码",
    "备注",
    "计划提测完成时间",
    "关联任务",
    "计划人天",
    "负责人",
    "测试人员名称",
    "需求人员名称",
    "产品",
    "创建时间",
    "创建人",
    "实际人天",
    "计划开发人天",
    "关联缺陷",
    "版本",
    "优先级",
    "绝对优先级",
    "关联用例",
    "计划主干测试完成时间",
    "修改时间",
    "验收标准",
    "详细说明（URL）",
    "立项状态",
    "迭代阶段",
    "迭代目标",
    "发布计划",
    "发布窗口",
    "Scrum团队",
    "产品组",
    "所属项目",
    "KSM或BUG编号",
    "类型",
    "所属云",
    "所属应用",
    "关联故事",
    "关联需求",
    "完成时间",
    "计划基线测试完成时间",
    "开发人员名称",
    "是否有升级注意事项",
    "变动类型",
]

# Canonical header -> DB column name
STORY_HEADER_COLUMN_MAP: dict[str, str] = {
    "序号": "sequence_no",
    "用户故事名称": "user_story_name",
    "用户故事标签": "user_story_tag",
    "状态": "status",
    "用户故事编码": "user_story_code",
    "备注": "remark",
    "计划提测完成时间": "plan_test_completion_time",
    "关联任务": "related_task_count",
    "计划人天": "planned_person_days",
    "负责人": "owner_names",
    "测试人员名称": "tester_names",
    "需求人员名称": "requirement_owner_names",
    "产品": "product",
    "创建时间": "created_time",
    "创建人": "created_by",
    "实际人天": "actual_person_days",
    "计划开发人天": "planned_dev_person_days",
    "关联缺陷": "related_defect_count",
    "版本": "version",
    "优先级": "priority",
    "绝对优先级": "absolute_priority",
    "关联用例": "related_case_count",
    "计划主干测试完成时间": "plan_trunk_test_completion_time",
    "修改时间": "modified_time",
    "验收标准": "acceptance_criteria",
    "详细说明（URL）": "detail_url",
    "立项状态": "project_status",
    "迭代阶段": "iteration_phase",
    "迭代目标": "iteration_goal",
    "发布计划": "release_plan",
    "发布窗口": "release_window",
    "Scrum团队": "scrum_team",
    "产品组": "product_group",
    "所属项目": "project_name",
    "KSM或BUG编号": "ksm_or_bug_no",
    "类型": "story_type",
    "所属云": "cloud_name",
    "所属应用": "application_name",
    "关联故事": "related_story",
    "关联需求": "related_requirement",
    "完成时间": "completed_time",
    "计划基线测试完成时间": "plan_baseline_test_completion_time",
    "开发人员名称": "developer_names",
    "是否有升级注意事项": "has_upgrade_notice",
    "变动类型": "change_type",
}

STORY_DB_COLUMNS: list[str] = [STORY_HEADER_COLUMN_MAP[header] for header in CANONICAL_STORY_HEADERS]

HEADER_ALIASES: dict[str, str] = {
    "详细说明(url)": "详细说明（URL）",
    "详细说明（url）": "详细说明（URL）",
    "scrum团队": "Scrum团队",
    "scrumteam": "Scrum团队",
    "ksm/bug编号": "KSM或BUG编号",
    "ksm或bug编号": "KSM或BUG编号",
}

NUMERIC_HEADERS = {
    "序号",
    "关联任务",
    "计划人天",
    "实际人天",
    "计划开发人天",
    "关联缺陷",
    "绝对优先级",
    "关联用例",
}

DATETIME_HEADERS = {
    "计划提测完成时间",
    "创建时间",
    "计划主干测试完成时间",
    "修改时间",
    "完成时间",
    "计划基线测试完成时间",
}


@dataclass
class StoryImportError:
    row: int
    reason: str


@dataclass
class StoryExcelImportResult:
    total_rows: int
    valid_rows: list[dict[str, Any]]
    errors: list[StoryImportError]
    missing_headers: list[str]
    unknown_headers: list[str]


class StoryExcelHeaderError(ValueError):
    def __init__(self, missing_headers: list[str], unknown_headers: list[str]) -> None:
        missing_text = "、".join(missing_headers) if missing_headers else "无"
        unknown_text = "、".join(unknown_headers) if unknown_headers else "无"
        super().__init__(f"故事 Excel 模板不匹配，缺失列: {missing_text}；未知列: {unknown_text}")
        self.missing_headers = missing_headers
        self.unknown_headers = unknown_headers


def normalize_story_header(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = text.replace("（", "(").replace("）", ")").replace("：", ":")
    text = re.sub(r"\s+", "", text)
    return text.lower()


def _resolve_canonical_header(value: Any) -> str:
    normalized = normalize_story_header(value)
    if not normalized:
        return ""
    if normalized in HEADER_ALIASES:
        return HEADER_ALIASES[normalized]
    for header in CANONICAL_STORY_HEADERS:
        if normalize_story_header(header) == normalized:
            return header
    return ""


def _parse_numeric(value: Any) -> float | None:
    text = normalize_text(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_datetime(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    text = normalize_text(value)
    if not text:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt in {"%Y-%m-%d", "%Y/%m/%d"}:
                return parsed.date().isoformat()
            return parsed.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return text


def _parse_story_cell(header: str, value: Any) -> Any:
    if header in NUMERIC_HEADERS:
        return _parse_numeric(value)
    if header in DATETIME_HEADERS:
        return _parse_datetime(value)
    text = normalize_text(value)
    return text or None


def import_story_excel(source: ExcelInput) -> StoryExcelImportResult:
    workbook = load_workbook(workbook_stream(source), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    unknown_headers: list[str] = []
    header_indexes: dict[str, int] = {}

    for idx, raw_header in enumerate(header_row):
        canonical = _resolve_canonical_header(raw_header)
        if not canonical:
            text = normalize_text(raw_header)
            if text:
                unknown_headers.append(text)
            continue
        if canonical not in header_indexes:
            header_indexes[canonical] = idx

    missing_headers = [header for header in CANONICAL_STORY_HEADERS if header not in header_indexes]
    if missing_headers:
        raise StoryExcelHeaderError(missing_headers=missing_headers, unknown_headers=unknown_headers)

    total_rows = 0
    valid_rows: list[dict[str, Any]] = []
    errors: list[StoryImportError] = []

    for row_idx, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        first_value = normalize_text(values[0]) if values else ""
        if first_value == "合计":
            break
        if not any(value not in (None, "") for value in values):
            continue

        total_rows += 1
        parsed_row: dict[str, Any] = {}
        for header in CANONICAL_STORY_HEADERS:
            value = values[header_indexes[header]] if header_indexes[header] < len(values) else None
            parsed_row[STORY_HEADER_COLUMN_MAP[header]] = _parse_story_cell(header, value)

        user_story_code = normalize_text(parsed_row.get("user_story_code"))
        if not user_story_code:
            errors.append(StoryImportError(row=row_idx, reason="缺少用户故事编码"))
            continue

        parsed_row["user_story_code"] = user_story_code
        valid_rows.append(parsed_row)

    return StoryExcelImportResult(
        total_rows=total_rows,
        valid_rows=valid_rows,
        errors=errors,
        missing_headers=missing_headers,
        unknown_headers=unknown_headers,
    )


def row_to_story_record(row: Mapping[str, Any]) -> StoryRecord:
    planned_person_days = float(row.get("planned_dev_person_days") or row.get("planned_person_days") or 0.0)
    actual_person_days = float(row.get("actual_person_days") or 0.0)
    owner_names = normalize_text(row.get("owner_names"))
    tester_names = normalize_text(row.get("tester_names"))
    developer_names = normalize_text(row.get("developer_names"))
    related_defect_count = row.get("related_defect_count")
    return StoryRecord(
        user_story_code=normalize_text(row.get("user_story_code")),
        name=normalize_text(row.get("user_story_name")),
        user_story_name=normalize_text(row.get("user_story_name")),
        sequence_no=row.get("sequence_no"),
        user_story_tag=normalize_text(row.get("user_story_tag")),
        status=normalize_text(row.get("status")),
        remark=normalize_text(row.get("remark")),
        plan_test_completion_time=normalize_text(row.get("plan_test_completion_time")),
        related_task_count=row.get("related_task_count"),
        owner=owner_names,
        owner_names=owner_names,
        tester=tester_names,
        tester_names=tester_names,
        requirement_owner_names=normalize_text(row.get("requirement_owner_names")),
        developers=split_names(row.get("developer_names") or row.get("owner_names")),
        developer_names=developer_names,
        priority=normalize_text(row.get("priority")) or "中",
        absolute_priority=row.get("absolute_priority"),
        plan_test_date=normalize_text(
            row.get("plan_test_completion_time") or row.get("plan_trunk_test_completion_time")
        ),
        planned_person_days=planned_person_days,
        planned_dev_person_days=float(row.get("planned_dev_person_days") or 0.0),
        actual_person_days=actual_person_days,
        related_case_count=row.get("related_case_count"),
        plan_trunk_test_completion_time=normalize_text(row.get("plan_trunk_test_completion_time")),
        modified_time=normalize_text(row.get("modified_time")),
        acceptance_criteria=normalize_text(row.get("acceptance_criteria")),
        detail_url=normalize_text(row.get("detail_url")),
        project_status=normalize_text(row.get("project_status")),
        iteration_phase=normalize_text(row.get("iteration_phase")),
        iteration_goal=normalize_text(row.get("iteration_goal")),
        release_plan=normalize_text(row.get("release_plan")),
        release_window=normalize_text(row.get("release_window")),
        scrum_team=normalize_text(row.get("scrum_team")),
        product_group=normalize_text(row.get("product_group")),
        project_name=normalize_text(row.get("project_name")),
        ksm_or_bug_no=normalize_text(row.get("ksm_or_bug_no")),
        story_type=normalize_text(row.get("story_type")),
        cloud_name=normalize_text(row.get("cloud_name")),
        application_name=normalize_text(row.get("application_name")),
        related_story=normalize_text(row.get("related_story")),
        related_requirement=normalize_text(row.get("related_requirement")),
        completed_time=normalize_text(row.get("completed_time")),
        plan_baseline_test_completion_time=normalize_text(row.get("plan_baseline_test_completion_time")),
        has_upgrade_notice=normalize_text(row.get("has_upgrade_notice")),
        change_type=normalize_text(row.get("change_type")),
        product=normalize_text(row.get("product")),
        version=normalize_text(row.get("version")),
        created_time=normalize_text(row.get("created_time")),
        created_by=normalize_text(row.get("created_by")),
        defects=int(float(related_defect_count or 0.0)),
        related_defect_count=related_defect_count if related_defect_count is not None else None,
        module_path=normalize_text(row.get("application_name")),
    )
